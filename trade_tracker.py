"""
trade_tracker.py
Builds trade_tracker.xlsx from trade_history.json.
Resolves open trades using eval logic against live yfinance data.
Called by scanner_combined.py and scanner_daily.py on every scheduled run.
"""
import json, os
import pandas as pd
import yfinance as yf
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

WIN_TARGET_PCT = 0.13
HOLD_WEEKS_WK  = 20
HOLD_DAYS_1D   = 60   # trading days

C_DARK    = '1A1A2E'; C_HDR = '2C3E7A'
C_WHITE   = 'FFFFFF'; C_GRAY = 'F5F5F5'; C_ALT = 'EEF2FF'
C_WIN     = 'D4EFDF'; C_LOSS = 'FADBD8'
C_NEUTRAL = 'FEF9E7'; C_OPEN = 'EBF5FB'
C_GREEN   = '1E8449'; C_RED  = 'C0392B'
C_BLACK   = '000000'; C_BLUE = '1A5276'

def _fill(c): return PatternFill('solid', start_color=c, fgColor=c)
def _ctr():   return Alignment(horizontal='center', vertical='center', wrap_text=True)
def _left():  return Alignment(horizontal='left',   vertical='center', wrap_text=True)
def _bdr():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def _font(size=10, bold=False, color=C_BLACK):
    return Font(name='Arial', size=size, bold=bold, color=color)

DATA_START = 6   # first data row (rows 1-5 are title/subtitle/spacer/legend/headers)

# ── Column definitions ─────────────────────────────────────────────────────────
# Each entry: (col_index, key_in_json_or_None, header, width, number_format, formula_or_None)
# formula uses {r}=row, {eE}=entry col letter, {fF}=stop col letter, etc.
COLS = [
    # col  json_key           header          w    fmt              formula
    (1,  'scan_date',       'Scan Date',     11, 'MM/DD/YYYY',    None),
    (2,  'ticker',          'Ticker',         8, '@',             None),
    (3,  'strategy',        'Strategy',      20, '@',             None),
    (4,  'interval',        'Interval',       8, '@',             None),
    (5,  'buy_price',       'Entry Price',    9, '$#,##0.00',     None),
    (6,  'stop_loss',       'Stop Loss',      9, '$#,##0.00',     None),
    (7,  None,              'Win Target',     9, '$#,##0.00',     '=IFERROR(E{r}*1.13,"")'),
    (8,  None,              'Shares',         7, '#,##0',         '=IFERROR(INT(10000/E{r}),"")'),
    (9,  None,              'Stop Dist %',    9, '0.00%',         '=IFERROR((E{r}-F{r})/E{r},"")'),
    (10, None,              'R/R',            7, '0.00"x"',       '=IFERROR((G{r}-E{r})/(E{r}-F{r}),"")'),
    (11, 'result',          'Result',         9, '@',             None),
    (12, 'date_bought',     'Date Bought',   11, 'MM/DD/YYYY',    None),
    (13, 'date_sold',       'Date Sold',     10, 'MM/DD/YYYY',    None),
    (14, 'exit_price',      'Exit Price',     9, '$#,##0.00',     None),
    (15, None,              'Hold (days)',   10, '#,##0',         '=IFERROR(M{r}-L{r},"")'),
    (16, None,              'Return %',       9, '0.00%',         '=IFERROR((N{r}-E{r})/E{r},"")'),
    (17, None,              'P&L ($)',       11, '$#,##0.00',     '=IFERROR((N{r}-E{r})*H{r},"")'),
    (18, None,              'Running P&L',   12, '$#,##0.00',     None),  # special cumulative
    (19, 'notes',           'Notes',         18, '@',             None),
]

def _cl(col_idx): return get_column_letter(col_idx)

# ── History I/O ────────────────────────────────────────────────────────────────
def load_history(path='trade_history.json'):
    if not os.path.exists(path):
        return []
    with open(path) as f:
        data = json.load(f)
    return data if isinstance(data, list) else []

def save_history(history, path='trade_history.json'):
    with open(path, 'w') as f:
        json.dump(history, f, indent=2, default=str)

# ── Add new signals at scan time ───────────────────────────────────────────────
def add_signals(history, new_signals):
    """
    new_signals: list of dicts with: scan_date, ticker, strategy, interval,
                 buy_price, stop_loss
    date_bought = scan_date (entry at trigger candle close = scan date)
    """
    for sig in new_signals:
        entry = float(sig.get('buy_price') or 0)
        stop  = float(sig.get('stop_loss') or 0)
        history.append({
            'scan_date':   str(sig['scan_date']),
            'ticker':      sig['ticker'],
            'strategy':    sig['strategy'],
            'interval':    sig['interval'],
            'buy_price':   round(entry, 4),
            'stop_loss':   round(stop,  4),
            'result':      'OPEN',
            'date_bought': str(sig['scan_date']),
            'date_sold':   '',
            'exit_price':  '',
            'notes':       '',
        })
    return history

# ── Resolve open trades ────────────────────────────────────────────────────────
def _expiry_date(date_bought_str, interval):
    """Returns the hold-period expiry date string."""
    start = pd.Timestamp(date_bought_str)
    if interval == '1wk':
        return str((start + pd.Timedelta(weeks=HOLD_WEEKS_WK)).date())
    else:
        # Estimate 60 trading days — use SPY calendar
        end_est = start + pd.Timedelta(days=100)
        try:
            spy = yf.download('SPY',
                              start=start.strftime('%Y-%m-%d'),
                              end=end_est.strftime('%Y-%m-%d'),
                              interval='1d', progress=False, auto_adjust=True)
            if len(spy) >= HOLD_DAYS_1D:
                return str(spy.index[HOLD_DAYS_1D - 1].date())
        except: pass
        return str((start + pd.Timedelta(days=84)).date())  # fallback ~60 trading days

def resolve_open_trades(history):
    """
    Walk all OPEN trades. For each, fetch daily OHLC from date_bought to today
    and apply the same eval logic as the backtest:
      bar HIGH >= win_target  → WIN  at win_target
      bar LOW  <= stop_loss   → LOSS at stop_loss
      hold period expired     → NEUTRAL at last close
    """
    today_str = str(date.today())

    for trade in history:
        if trade.get('result') != 'OPEN':
            continue
        entry = float(trade.get('buy_price') or 0)
        stop  = float(trade.get('stop_loss') or 0)
        if entry <= 0:
            continue

        win_target  = entry * (1 + WIN_TARGET_PCT)
        date_bought = trade.get('date_bought') or trade.get('scan_date')
        interval    = trade.get('interval', '1wk')
        expiry_str  = _expiry_date(date_bought, interval)

        if date_bought > today_str:
            continue

        try:
            df = yf.download(
                trade['ticker'],
                start=date_bought,
                end=(pd.Timestamp(today_str) + pd.Timedelta(days=1)).strftime('%Y-%m-%d'),
                interval='1d', progress=False, auto_adjust=True)
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            if df is None or df.empty:
                continue
        except Exception as e:
            print(f'  resolve: fetch failed {trade["ticker"]}: {e}')
            continue

        result = 'OPEN'; exit_price = None; exit_date = None

        for idx in range(len(df)):
            bar_date = str(df.index[idx].date())
            if bar_date > expiry_str:
                break
            bar_high = float(df['High'].iloc[idx])
            bar_low  = float(df['Low'].iloc[idx])
            if bar_high >= win_target:
                result = 'WIN';  exit_price = round(win_target, 4); exit_date = bar_date; break
            if bar_low <= stop:
                result = 'LOSS'; exit_price = round(stop, 4);       exit_date = bar_date; break

        if result == 'OPEN' and today_str >= expiry_str:
            exp_bars = df[df.index <= pd.Timestamp(expiry_str)]
            if not exp_bars.empty:
                result     = 'NEUTRAL'
                exit_price = round(float(exp_bars['Close'].iloc[-1]), 4)
                exit_date  = str(exp_bars.index[-1].date())

        if result != 'OPEN':
            trade['result']     = result
            trade['date_sold']  = exit_date
            trade['exit_price'] = exit_price
            print(f'  Resolved {trade["ticker"]} ({interval}): '
                  f'{result} @ ${exit_price} on {exit_date}')

    return history

# ── Build Excel ────────────────────────────────────────────────────────────────
def build_excel(history, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Trade Log'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f'A{DATA_START}'

    ncols    = len(COLS)
    last_col = _cl(ncols)

    # Row 1 — Title
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'STOCK SCANNER — PAPER TRADE TRACKER'
    ws['A1'].font      = _font(13, True, C_WHITE)
    ws['A1'].fill      = _fill(C_DARK)
    ws['A1'].alignment = _ctr()
    ws.row_dimensions[1].height = 26

    # Row 2 — Subtitle
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = (f'Entry = close of BB trigger candle  |  Stop = low of trigger candle  |  '
                f'Win target = {int(WIN_TARGET_PCT*100)}%  |  $10,000/trade  |  '
                f'Hold: {HOLD_WEEKS_WK}wk (1wk) / {HOLD_DAYS_1D} trading days (1d)  |  '
                f'Results resolved automatically on each scan run')
    ws['A2'].font      = _font(8, False, 'AAAAAA')
    ws['A2'].fill      = _fill(C_DARK)
    ws['A2'].alignment = _ctr()
    ws.row_dimensions[2].height = 15

    # Row 3 — Spacer
    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'].fill = _fill(C_DARK)
    ws.row_dimensions[3].height = 5

    # Row 4 — Legend
    ws.merge_cells(f'A4:{last_col}4')
    ws['A4'] = ('WIN = target hit within hold    LOSS = stop loss hit    '
                'NEUTRAL = hold expired, no trigger    OPEN = still active')
    ws['A4'].font      = _font(9, False, '333333')
    ws['A4'].fill      = _fill('F0F4FF')
    ws['A4'].alignment = _ctr()
    ws.row_dimensions[4].height = 15

    # Row 5 — Column headers
    for col_idx, key, hdr, width, fmt, formula in COLS:
        cell = ws.cell(5, col_idx)
        cell.value     = hdr
        cell.font      = _font(9, True, C_WHITE)
        cell.fill      = _fill(C_HDR)
        cell.alignment = _ctr()
        cell.border    = _bdr()
        ws.column_dimensions[_cl(col_idx)].width = width
    ws.row_dimensions[5].height = 28

    # Data rows
    for r_off, trade in enumerate(history):
        r        = DATA_START + r_off
        row_fill = _fill(C_ALT) if r_off % 2 == 0 else _fill(C_WHITE)

        for col_idx, key, hdr, width, fmt, formula in COLS:
            cell = ws.cell(r, col_idx)
            cell.fill          = row_fill
            cell.border        = _bdr()
            cell.alignment     = _ctr()
            cell.number_format = fmt
            cell.font          = _font(10)

            if col_idx == 18:
                # Running P&L — cumulative
                q_col = _cl(17)
                if r == DATA_START:
                    cell.value = f'=IFERROR({q_col}{r},"")'
                else:
                    cell.value = f'=IFERROR({q_col}{r}+{_cl(18)}{r-1},"")'
            elif formula:
                cell.value = formula.replace('{r}', str(r))
            elif key:
                val = trade.get(key, '')
                if val is not None and val != '':
                    cell.value = val

        ws.row_dimensions[r].height = 18

    # Conditional formatting on Result col (11)
    if history:
        d1 = DATA_START; d2 = DATA_START + len(history) - 1
        kr = f'{_cl(11)}{d1}:{_cl(11)}{d2}'
        ws.conditional_formatting.add(kr, CellIsRule('equal', ['"WIN"'],
            fill=_fill(C_WIN),     font=Font(bold=True, color=C_GREEN)))
        ws.conditional_formatting.add(kr, CellIsRule('equal', ['"LOSS"'],
            fill=_fill(C_LOSS),    font=Font(bold=True, color=C_RED)))
        ws.conditional_formatting.add(kr, CellIsRule('equal', ['"NEUTRAL"'],
            fill=_fill(C_NEUTRAL)))
        ws.conditional_formatting.add(kr, CellIsRule('equal', ['"OPEN"'],
            fill=_fill(C_OPEN),    font=Font(bold=True, color='1A6B9A')))
        for rng in [f'{_cl(17)}{d1}:{_cl(17)}{d2}', f'{_cl(18)}{d1}:{_cl(18)}{d2}']:
            ws.conditional_formatting.add(rng, CellIsRule('greaterThan', ['0'],
                font=Font(bold=True, color=C_GREEN)))
            ws.conditional_formatting.add(rng, CellIsRule('lessThan', ['0'],
                font=Font(bold=True, color=C_RED)))

    # Summary row
    if history:
        sr  = DATA_START + len(history) + 1
        d1  = DATA_START; d2 = DATA_START + len(history) - 1
        ws.merge_cells(f'A{sr}:{_cl(10)}{sr}')
        ws[f'A{sr}']        = 'SUMMARY'
        ws[f'A{sr}'].font   = _font(9, True, C_WHITE)
        ws[f'A{sr}'].fill   = _fill(C_HDR)
        ws[f'A{sr}'].alignment = _ctr()

        summary = {
            _cl(11): (f'=COUNTIF({_cl(11)}{d1}:{_cl(11)}{d2},"WIN")&" W  |  "'
                      f'&COUNTIF({_cl(11)}{d1}:{_cl(11)}{d2},"LOSS")&" L  |  "'
                      f'&COUNTIF({_cl(11)}{d1}:{_cl(11)}{d2},"NEUTRAL")&" N  |  "'
                      f'&COUNTIF({_cl(11)}{d1}:{_cl(11)}{d2},"OPEN")&" Open"'),
            _cl(16): (f'=IFERROR(AVERAGEIF({_cl(11)}{d1}:{_cl(11)}{d2},'
                      f'"<>OPEN",{_cl(16)}{d1}:{_cl(16)}{d2}),"")'),
            _cl(17): f'=IFERROR(SUM({_cl(17)}{d1}:{_cl(17)}{d2}),"")',
            _cl(18): f'=IFERROR({_cl(18)}{d2},"")',
        }
        fmts = {_cl(16): '0.00%', _cl(17): '$#,##0.00', _cl(18): '$#,##0.00'}
        for col_letter, formula in summary.items():
            c = ws[f'{col_letter}{sr}']
            c.value = formula; c.font = _font(9, True, C_WHITE)
            c.fill = _fill(C_HDR); c.alignment = _ctr(); c.border = _bdr()
            if col_letter in fmts:
                c.number_format = fmts[col_letter]
        ws.row_dimensions[sr].height = 18

    # ── Strategy Stats sheet ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Strategy Stats')
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:M1')
    ws2['A1'] = 'STRATEGY STATS — Historical Backtest Reference'
    ws2['A1'].font = _font(11, True, C_WHITE); ws2['A1'].fill = _fill(C_DARK)
    ws2['A1'].alignment = _ctr(); ws2.row_dimensions[1].height = 22

    s_hdrs  = ['Strategy','Interval','Hold','Signals\n(15yr)','~/mo',
               'Win %','Loss %','Avg Win','Avg Loss','Exp Value','Avg Hold','Total P&L','ROI']
    s_widths= [22,9,10,12,7,8,8,9,9,10,10,13,8]
    for col,(h,w) in enumerate(zip(s_hdrs,s_widths),1):
        c = ws2.cell(3,col); c.value=h
        c.font=_font(9,True,C_WHITE); c.fill=_fill(C_HDR)
        c.alignment=_ctr(); c.border=_bdr()
        ws2.column_dimensions[_cl(col)].width=w
    ws2.row_dimensions[3].height=28

    rows=[
        ('1WK — Tier 1 Ultra','1wk','20 wks',48, 0.3,0.960,0.000,0.130,0.000,0.1248,'5.1w', 62498, 0.125),
        ('1WK — Tier 2 High', '1wk','20 wks',106,0.6,0.944,0.000,0.130,0.000,None,  None,   None,  0.125),
        ('1D  — Tier 1 Ultra','1d', '60 days',151,0.8,0.858,None, None, None, None,  '23.2d',198375,0.131),
        ('1D  — Tier 2 High', '1d', '60 days',313,1.7,0.880,None, None, None, None,  '20.1d',418341,0.134),
    ]
    pct_c={6,7,8,9,10,13}
    for r_off,row in enumerate(rows):
        r=4+r_off; rf=_fill(C_WHITE) if r_off%2==0 else _fill(C_ALT)
        for col,val in enumerate(row,1):
            c=ws2.cell(r,col); c.fill=rf; c.border=_bdr(); c.alignment=_ctr()
            if val is None:
                c.value='—'; c.font=_font(10,False,'AAAAAA')
            else:
                c.value=val; c.font=_font(10,False,C_BLUE)
                if col in pct_c and isinstance(val,float): c.number_format='0.0%'
                if col==12 and isinstance(val,(int,float)): c.number_format='$#,##0'
        ws2.row_dimensions[r].height=18

    nr=4+len(rows)+1
    ws2.merge_cells(f'A{nr}:M{nr}')
    ws2[f'A{nr}']='Update after re-running backtests via GitHub Actions → Backtests workflow.'
    ws2[f'A{nr}'].font=_font(9,False,'777777'); ws2[f'A{nr}'].fill=_fill(C_GRAY)
    ws2[f'A{nr}'].alignment=_left(); ws2.row_dimensions[nr].height=16

    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    h = load_history()
    print(f'Loaded {len(h)} trades')
    h = resolve_open_trades(h)
    save_history(h)
    path = build_excel(h, '/tmp/trade_tracker_test.xlsx')
    print(f'Built: {path}')
